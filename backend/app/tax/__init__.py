"""Tax prep: prior-year ingest, doc sorter, book-to-tax, lead sheet."""

from __future__ import annotations

import json
import re
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app import db
from app.ledger.service import get_transactions, trial_balance
from app.money import money
from app.paths import DICTIONARIES
from app.tax import constants_2025 as C


def load_book_to_tax_map(entity_type: str) -> list[dict[str, Any]]:
    path = DICTIONARIES / "book_to_tax.json"
    data = json.loads(path.read_text()) if path.exists() else {}
    return data.get(entity_type, data.get("sole_prop", []))


def ingest_prior_year(client_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    tax_year = int(payload.get("tax_year") or payload.get("year"))
    db.save_prior_year(client_id, tax_year, payload)

    # Generate checklist from forms present / expected
    forms = payload.get("forms") or payload.get("documents") or []
    defaults = [
        ("W2", "Form W-2 Wage and Tax Statement"),
        ("1099NEC", "Form 1099-NEC Nonemployee Compensation"),
        ("1099INT", "Form 1099-INT Interest Income"),
        ("1099DIV", "Form 1099-DIV Dividends"),
        ("1099MISC", "Form 1099-MISC"),
        ("K1", "Schedule K-1"),
        ("1098", "Form 1098 Mortgage Interest"),
        ("PRIOR_RETURN", "Prior-year tax return PDF"),
    ]
    seen = {str(f.get("type") or f).upper().replace("-", "").replace("_", "") for f in forms if isinstance(f, dict) or isinstance(f, str)}
    for doc_type, label in defaults:
        # always create core checklist; mark required based on prior year
        required = doc_type in seen or doc_type in ("PRIOR_RETURN", "W2", "1099NEC")
        db.upsert_checklist_item(client_id, tax_year + 1 if "next" not in payload else tax_year, doc_type, label, required=True)

    # Also seed checklist for current prep year = tax_year + 1 typically
    prep_year = int(payload.get("prep_year") or (tax_year + 1))
    for doc_type, label in defaults:
        db.upsert_checklist_item(client_id, prep_year, doc_type, label, required=True)

    questionnaire = generate_questionnaire(payload)
    return {
        "tax_year": tax_year,
        "checklist": db.list_checklist(client_id, prep_year),
        "questionnaire": questionnaire,
        "disclaimer": C.DISCLAIMER,
    }


def generate_questionnaire(payload: dict[str, Any]) -> list[dict[str, str]]:
    qs = [
        {"id": "any_new_w2", "prompt": "Did you receive any new W-2s this year?"},
        {"id": "contractors", "prompt": "Did you pay any contractors $600+ (need 1099-NEC)?"},
        {"id": "home_office", "prompt": "Did you use a home office exclusively for business?"},
        {"id": "vehicle", "prompt": "Did you use a vehicle for business? Miles?"},
        {"id": "augusta", "prompt": "Did the business rent your home for meetings (Augusta Rule)?"},
        {"id": "equipment", "prompt": "Did you purchase equipment over $2,500 (Section 179 candidate)?"},
    ]
    if payload.get("entity_type") == "s_corp" or "S-Corp" in str(payload.get("entity") or ""):
        qs.append({"id": "reasonable_comp", "prompt": "Confirm owner salary paid via payroll and hours worked."})
    schedules = payload.get("schedules") or []
    if any("C" in str(s) for s in schedules):
        qs.append({"id": "cogs", "prompt": "Any inventory / COGS changes vs prior Schedule C?"})
    return qs


DOC_TYPE_PATTERNS = [
    (re.compile(r"w-?2", re.I), "W2"),
    (re.compile(r"1099-?nec|nonemployee", re.I), "1099NEC"),
    (re.compile(r"1099-?int|interest", re.I), "1099INT"),
    (re.compile(r"1099-?div|dividend", re.I), "1099DIV"),
    (re.compile(r"1099-?misc", re.I), "1099MISC"),
    (re.compile(r"k-?1", re.I), "K1"),
    (re.compile(r"1098", re.I), "1098"),
    (re.compile(r"1040|tax.?return", re.I), "PRIOR_RETURN"),
]


def detect_doc_type(filename: str, content: bytes | None = None) -> str | None:
    name = filename.lower()
    for pattern, doc_type in DOC_TYPE_PATTERNS:
        if pattern.search(name):
            return doc_type
    if content:
        text = content[:5000].decode("latin-1", errors="ignore")
        for pattern, doc_type in DOC_TYPE_PATTERNS:
            if pattern.search(text):
                return doc_type
    return None


def sort_documents(client_id: str, tax_year: int, files: list[tuple[str, bytes]]) -> list[dict[str, Any]]:
    results = []
    for filename, content in files:
        doc_type = detect_doc_type(filename, content)
        checked = False
        if doc_type:
            checked = db.mark_checklist_received(client_id, tax_year, doc_type, file_path=filename)
        results.append(
            {
                "filename": filename,
                "detected_type": doc_type,
                "checklist_updated": checked,
            }
        )
    return results


def book_to_tax_grid(client_id: str, path: str, entity_type: str) -> dict[str, Any]:
    mapping = load_book_to_tax_map(entity_type)
    balances = {r["account"]: money(r["balance"]) for r in trial_balance(path)}

    lines: list[dict[str, Any]] = []
    for m in mapping:
        accounts = m.get("accounts") or [m.get("account")]
        accounts = [a for a in accounts if a]
        book_amt = money(sum((balances.get(a, money(0)) for a in accounts), money(0)))
        # Expenses in beancount trial balance are typically positive as debits;
        # for P&L tax lines we want absolute expense amounts
        if any(a.startswith("Expenses:") for a in accounts):
            book_amt = abs(book_amt)
        if any(a.startswith("Income:") for a in accounts):
            book_amt = abs(book_amt)

        adjustment = money(0)
        tax_amt = book_amt
        note = ""
        if m.get("meals_limitation"):
            deductible = money(book_amt * C.MEALS_DEDUCTIBLE_PCT)
            adjustment = money(book_amt - deductible)
            tax_amt = deductible
            note = f"{int(C.MEALS_DEDUCTIBLE_PCT * 100)}% M&E limitation applied"

        lines.append(
            {
                "tax_line": m["tax_line"],
                "description": m["description"],
                "accounts": accounts,
                "book_amount": str(book_amt),
                "adjustment": str(adjustment),
                "tax_amount": str(tax_amt),
                "note": note,
                "form": m.get("form", "Schedule C"),
            }
        )

    return {
        "entity_type": entity_type,
        "lines": lines,
        "disclaimer": C.DISCLAIMER,
        "meals_pct": str(C.MEALS_DEDUCTIBLE_PCT),
    }


def generate_lead_sheet(client_id: str, path: str, entity_type: str) -> bytes:
    grid = book_to_tax_grid(client_id, path, entity_type)
    txs = get_transactions(path)

    wb = Workbook()
    summary = wb.active
    summary.title = "Tax Line Summary"
    summary.append(
        [
            "Tax Line Code",
            "Description",
            "Form",
            "Book Amount",
            "M-1 / BTD Adjustment",
            "Tax Amount",
            "Source Documents",
            "Notes",
        ]
    )
    for line in grid["lines"]:
        summary.append(
            [
                line["tax_line"],
                line["description"],
                line["form"],
                line["book_amount"],
                line["adjustment"],
                line["tax_amount"],
                "",
                line["note"],
            ]
        )

    detail = wb.create_sheet("Transaction Detail")
    detail.append(["Date", "Payee", "Narration", "Account", "Amount", "Tax Line"])
    acct_to_line = {}
    for line in grid["lines"]:
        for a in line["accounts"]:
            acct_to_line[a] = line["tax_line"]

    for tx in txs:
        for p in tx["postings"]:
            detail.append(
                [
                    tx["date"],
                    tx.get("payee"),
                    tx.get("narration"),
                    p["account"],
                    p["amount"],
                    acct_to_line.get(p["account"], ""),
                ]
            )

    meta = wb.create_sheet("Meta")
    meta.append(["Client", client_id])
    meta.append(["Entity Type", entity_type])
    meta.append(["Disclaimer", C.DISCLAIMER])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
